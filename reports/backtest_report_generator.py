# reports/backtest_report_generator.py
"""
Professional Backtest Report Generator
Creates comprehensive, professional-grade reports in multiple formats (HTML, PDF, Excel)
"""

import logging
import json
import os
from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
import pandas as pd
import numpy as np
from jinja2 import Template, Environment, FileSystemLoader
import base64
import io

# Try to import optional dependencies
try:
    import weasyprint
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
    from openpyxl.drawing.image import Image as OpenpyxlImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import matplotlib.pyplot as plt
    import seaborn as sns
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

logger = logging.getLogger(__name__)


class BacktestReportGenerator:
    """
    Professional report generator for backtest results
    """
    
    def __init__(self, config: Dict[str, Any] = None):
        self.config = config or {}
        
        # Report settings
        self.output_dir = self.config.get('output_dir', 'reports/generated')
        self.template_dir = self.config.get('template_dir', 'reports/templates')
        self.company_name = self.config.get('company_name', 'Trading Analytics')
        self.report_title = self.config.get('report_title', 'Algorithmic Trading Strategy Analysis')
        self.author = self.config.get('author', 'Strategy Research Team')
        
        # Create directories
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
        Path(self.template_dir).mkdir(parents=True, exist_ok=True)
        
        # Initialize Jinja2 environment
        self.jinja_env = Environment(
            loader=FileSystemLoader(self.template_dir),
            autoescape=True
        )
        
        # Report styling
        self.colors = {
            'primary': '#2E86AB',
            'secondary': '#A23B72',
            'success': '#28A745',
            'danger': '#DC3545',
            'warning': '#FFC107',
            'info': '#17A2B8',
            'light': '#F8F9FA',
            'dark': '#343A40'
        }
        
        logger.info("BacktestReportGenerator initialized")
    
    def generate_comprehensive_report(self, 
                                    backtest_results: Dict[str, Any],
                                    analysis_results: Dict[str, Any] = None,
                                    optimization_results: Dict[str, Any] = None,
                                    market_data: pd.DataFrame = None,
                                    report_format: str = 'html') -> str:
        """
        Generate comprehensive professional report
        """
        try:
            logger.info(f"Generating comprehensive report in {report_format} format")
            
            # Prepare report data
            report_data = self._prepare_report_data(
                backtest_results, analysis_results, optimization_results, market_data
            )
            
            # Generate report based on format
            if report_format.lower() == 'html':
                return self._generate_html_report(report_data)
            elif report_format.lower() == 'pdf':
                return self._generate_pdf_report(report_data)
            elif report_format.lower() == 'excel':
                return self._generate_excel_report(report_data)
            elif report_format.lower() == 'all':
                # Generate all formats
                html_path = self._generate_html_report(report_data)
                pdf_path = self._generate_pdf_report(report_data)
                excel_path = self._generate_excel_report(report_data)
                return {'html': html_path, 'pdf': pdf_path, 'excel': excel_path}
            else:
                raise ValueError(f"Unsupported report format: {report_format}")
                
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            raise
    
    def _prepare_report_data(self, 
                           backtest_results: Dict[str, Any],
                           analysis_results: Dict[str, Any] = None,
                           optimization_results: Dict[str, Any] = None,
                           market_data: pd.DataFrame = None) -> Dict[str, Any]:
        """Prepare and structure data for report generation"""
        try:
            # Extract basic information
            strategy_name = backtest_results.get('strategy', 'Unknown Strategy')
            symbol = backtest_results.get('symbol', 'Unknown Symbol')
            start_date = backtest_results.get('start_date', 'Unknown')
            end_date = backtest_results.get('end_date', 'Unknown')
            initial_capital = backtest_results.get('initial_capital', 0)
            final_capital = backtest_results.get('final_capital', 0)
            
            # Extract trades and equity curve
            trades_df = pd.DataFrame(backtest_results.get('trades', []))
            equity_curve_df = pd.DataFrame(backtest_results.get('equity_curve', []))
            
            # Performance metrics
            performance_metrics = {}
            risk_metrics = {}
            trade_metrics = {}
            
            if analysis_results:
                performance_metrics = analysis_results.get('performance_metrics', {})
                risk_metrics = analysis_results.get('risk_analysis', {})
                trade_analysis = analysis_results.get('trade_analysis', {})
                trade_metrics = trade_analysis.get('return_distribution', {}) if trade_analysis else {}
            
            # Calculate additional metrics if not available
            if not performance_metrics and not trades_df.empty:
                performance_metrics = self._calculate_basic_metrics(
                    trades_df, equity_curve_df, initial_capital, final_capital
                )
            
            # Market comparison
            market_comparison = {}
            if market_data is not None and not market_data.empty:
                market_comparison = self._calculate_market_comparison(
                    equity_curve_df, market_data
                )
            
            # Optimization insights
            optimization_insights = {}
            if optimization_results:
                optimization_insights = self._extract_optimization_insights(optimization_results)
            
            # Trade statistics
            trade_statistics = {}
            if not trades_df.empty:
                trade_statistics = self._calculate_trade_statistics(trades_df)
            
            # Risk analysis
            risk_analysis = {}
            if not equity_curve_df.empty:
                risk_analysis = self._calculate_risk_analysis(equity_curve_df)
            
            # Monthly/Quarterly breakdown
            period_analysis = {}
            if not trades_df.empty:
                period_analysis = self._calculate_period_analysis(trades_df)
            
            # Prepare final report data
            report_data = {
                'metadata': {
                    'report_title': self.report_title,
                    'company_name': self.company_name,
                    'author': self.author,
                    'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'strategy_name': strategy_name,
                    'symbol': symbol,
                    'test_period': f"{start_date} to {end_date}",
                    'initial_capital': initial_capital,
                    'final_capital': final_capital
                },
                'executive_summary': self._create_executive_summary(
                    performance_metrics, risk_metrics, trade_statistics
                ),
                'performance_metrics': performance_metrics,
                'risk_metrics': risk_metrics,
                'trade_metrics': trade_metrics,
                'trade_statistics': trade_statistics,
                'risk_analysis': risk_analysis,
                'market_comparison': market_comparison,
                'period_analysis': period_analysis,
                'optimization_insights': optimization_insights,
                'trades_data': trades_df.to_dict('records') if not trades_df.empty else [],
                'equity_curve_data': equity_curve_df.to_dict('records') if not equity_curve_df.empty else [],
                'recommendations': self._generate_recommendations(
                    performance_metrics, risk_metrics, optimization_insights
                ),
                'charts': self._prepare_chart_data(trades_df, equity_curve_df, market_data)
            }
            
            return report_data
            
        except Exception as e:
            logger.error(f"Error preparing report data: {e}")
            return {}
    
    def _calculate_basic_metrics(self, trades_df: pd.DataFrame, equity_curve_df: pd.DataFrame,
                               initial_capital: float, final_capital: float) -> Dict[str, Any]:
        """Calculate basic performance metrics"""
        try:
            metrics = {}
            
            # Basic returns
            total_return = (final_capital - initial_capital) / initial_capital if initial_capital > 0 else 0
            metrics['total_return'] = total_return
            metrics['total_return_pct'] = total_return * 100
            
            # Trade metrics
            if not trades_df.empty:
                winning_trades = trades_df[trades_df['net_pnl'] > 0]
                losing_trades = trades_df[trades_df['net_pnl'] <= 0]
                
                metrics['total_trades'] = len(trades_df)
                metrics['winning_trades'] = len(winning_trades)
                metrics['losing_trades'] = len(losing_trades)
                metrics['win_rate'] = len(winning_trades) / len(trades_df) if len(trades_df) > 0 else 0
                metrics['win_rate_pct'] = metrics['win_rate'] * 100
                
                # P&L metrics
                gross_profit = winning_trades['net_pnl'].sum() if not winning_trades.empty else 0
                gross_loss = abs(losing_trades['net_pnl'].sum()) if not losing_trades.empty else 0
                metrics['gross_profit'] = gross_profit
                metrics['gross_loss'] = gross_loss
                metrics['profit_factor'] = gross_profit / gross_loss if gross_loss > 0 else float('inf')
                
                metrics['avg_win'] = winning_trades['net_pnl'].mean() if not winning_trades.empty else 0
                metrics['avg_loss'] = losing_trades['net_pnl'].mean() if not losing_trades.empty else 0
                metrics['largest_win'] = trades_df['net_pnl'].max()
                metrics['largest_loss'] = trades_df['net_pnl'].min()
            
            # Equity curve metrics
            if not equity_curve_df.empty:
                equity_values = equity_curve_df['equity']
                
                # Max drawdown
                peak = equity_values.expanding().max()
                drawdown = (equity_values - peak) / peak
                metrics['max_drawdown'] = abs(drawdown.min())
                metrics['max_drawdown_pct'] = abs(drawdown.min()) * 100
                
                # Returns and volatility
                returns = equity_values.pct_change().dropna()
                if len(returns) > 1:
                    metrics['volatility'] = returns.std() * np.sqrt(252)  # Annualized
                    metrics['volatility_pct'] = metrics['volatility'] * 100
                    
                    # Sharpe ratio (assuming 2% risk-free rate)
                    excess_return = returns.mean() * 252 - 0.02
                    metrics['sharpe_ratio'] = excess_return / metrics['volatility'] if metrics['volatility'] > 0 else 0
            
            return metrics
            
        except Exception as e:
            logger.error(f"Error calculating basic metrics: {e}")
            return {}
    
    def _calculate_market_comparison(self, equity_curve_df: pd.DataFrame, 
                                   market_data: pd.DataFrame) -> Dict[str, Any]:
        """Calculate strategy performance vs market"""
        try:
            comparison = {}
            
            if equity_curve_df.empty or market_data.empty:
                return comparison
            
            # Convert timestamps
            equity_curve_df['timestamp'] = pd.to_datetime(equity_curve_df['timestamp'])
            
            if 'timestamp' not in market_data.columns:
                market_data = market_data.reset_index()
                if market_data.index.name == 'timestamp':
                    market_data['timestamp'] = market_data.index
            
            market_data['timestamp'] = pd.to_datetime(market_data['timestamp'])
            
            # Align data
            equity_curve_df = equity_curve_df.set_index('timestamp')
            market_data = market_data.set_index('timestamp')
            
            # Calculate returns
            strategy_returns = equity_curve_df['equity'].pct_change().dropna()
            market_returns = market_data['close'].pct_change().dropna() if 'close' in market_data.columns else pd.Series()
            
            if len(strategy_returns) > 0 and len(market_returns) > 0:
                # Find common dates
                common_dates = strategy_returns.index.intersection(market_returns.index)
                
                if len(common_dates) > 1:
                    strategy_ret = strategy_returns.loc[common_dates]
                    market_ret = market_returns.loc[common_dates]
                    
                    # Calculate metrics
                    comparison['correlation'] = strategy_ret.corr(market_ret)
                    comparison['beta'] = strategy_ret.cov(market_ret) / market_ret.var() if market_ret.var() > 0 else 0
                    
                    # Cumulative returns
                    strategy_cumret = (1 + strategy_ret).cumprod().iloc[-1] - 1
                    market_cumret = (1 + market_ret).cumprod().iloc[-1] - 1
                    
                    comparison['strategy_total_return'] = strategy_cumret
                    comparison['market_total_return'] = market_cumret
                    comparison['excess_return'] = strategy_cumret - market_cumret
                    
                    # Volatility comparison
                    comparison['strategy_volatility'] = strategy_ret.std() * np.sqrt(252)
                    comparison['market_volatility'] = market_ret.std() * np.sqrt(252)
                    
                    # Information ratio
                    active_return = strategy_ret - market_ret
                    tracking_error = active_return.std() * np.sqrt(252)
                    comparison['information_ratio'] = (active_return.mean() * 252) / tracking_error if tracking_error > 0 else 0
            
            return comparison
            
        except Exception as e:
            logger.error(f"Error calculating market comparison: {e}")
            return {}
    
    def _extract_optimization_insights(self, optimization_results: Dict[str, Any]) -> Dict[str, Any]:
        """Extract key insights from optimization results"""
        try:
            insights = {}
            
            if 'best_result' in optimization_results and optimization_results['best_result']:
                best_result = optimization_results['best_result']
                insights['best_parameters'] = best_result.get('parameters', {})
                insights['best_objective_value'] = best_result.get('objective_value', 0)
            
            if 'parameter_importance' in optimization_results:
                insights['parameter_importance'] = optimization_results['parameter_importance']
            
            if 'robust_parameters' in optimization_results:
                insights['robust_parameters'] = optimization_results['robust_parameters']
            
            if 'optimization_duration' in optimization_results:
                insights['optimization_duration'] = optimization_results['optimization_duration']
            
            if 'total_combinations' in optimization_results:
                insights['combinations_tested'] = optimization_results['total_combinations']
            
            return insights
            
        except Exception as e:
            logger.error(f"Error extracting optimization insights: {e}")
            return {}
    
    def _calculate_trade_statistics(self, trades_df: pd.DataFrame) -> Dict[str, Any]:
        """Calculate detailed trade statistics"""
        try:
            stats = {}
            
            if trades_df.empty:
                return stats
            
            # Convert timestamp
            trades_df['timestamp'] = pd.to_datetime(trades_df['timestamp'])
            
            # Basic stats
            stats['total_trades'] = len(trades_df)
            stats['first_trade'] = trades_df['timestamp'].min().strftime('%Y-%m-%d %H:%M')
            stats['last_trade'] = trades_df['timestamp'].max().strftime('%Y-%m-%d %H:%M')
            
            # P&L statistics
            pnl_values = trades_df['net_pnl']
            stats['total_pnl'] = pnl_values.sum()
            stats['avg_pnl_per_trade'] = pnl_values.mean()
            stats['median_pnl'] = pnl_values.median()
            stats['std_pnl'] = pnl_values.std()
            
            # Win/Loss analysis
            winning_trades = trades_df[trades_df['net_pnl'] > 0]
            losing_trades = trades_df[trades_df['net_pnl'] <= 0]
            
            stats['winning_trades_count'] = len(winning_trades)
            stats['losing_trades_count'] = len(losing_trades)
            stats['win_rate'] = len(winning_trades) / len(trades_df) * 100
            
            if not winning_trades.empty:
                stats['avg_winning_trade'] = winning_trades['net_pnl'].mean()
                stats['largest_win'] = winning_trades['net_pnl'].max()
            
            if not losing_trades.empty:
                stats['avg_losing_trade'] = losing_trades['net_pnl'].mean()
                stats['largest_loss'] = losing_trades['net_pnl'].min()
            
            # Consecutive wins/losses
            trades_sorted = trades_df.sort_values('timestamp')
            win_lose_sequence = (trades_sorted['net_pnl'] > 0).astype(int)
            
            # Calculate streaks
            stats['max_consecutive_wins'] = self._calculate_max_streak(win_lose_sequence, 1)
            stats['max_consecutive_losses'] = self._calculate_max_streak(win_lose_sequence, 0)
            
            # Trade frequency
            trading_days = (trades_df['timestamp'].max() - trades_df['timestamp'].min()).days
            stats['trades_per_day'] = len(trades_df) / trading_days if trading_days > 0 else 0
            
            # Trade duration analysis (if entry/exit times available)
            if 'entry_time' in trades_df.columns and 'exit_time' in trades_df.columns:
                entry_times = pd.to_datetime(trades_df['entry_time'])
                exit_times = pd.to_datetime(trades_df['exit_time'])
                durations = (exit_times - entry_times).dt.total_seconds() / 3600  # Hours
                
                stats['avg_trade_duration_hours'] = durations.mean()
                stats['median_trade_duration_hours'] = durations.median()
                stats['min_trade_duration_hours'] = durations.min()
                stats['max_trade_duration_hours'] = durations.max()
            
            return stats
            
        except Exception as e:
            logger.error(f"Error calculating trade statistics: {e}")
            return {}
    
    def _calculate_max_streak(self, sequence: pd.Series, value: int) -> int:
        """Calculate maximum consecutive streak of a specific value"""
        try:
            max_streak = 0
            current_streak = 0
            
            for val in sequence:
                if val == value:
                    current_streak += 1
                    max_streak = max(max_streak, current_streak)
                else:
                    current_streak = 0
            
            return max_streak
            
        except Exception as e:
            logger.error(f"Error calculating max streak: {e}")
            return 0
    
    def _calculate_risk_analysis(self, equity_curve_df: pd.DataFrame) -> Dict[str, Any]:
        """Calculate detailed risk analysis"""
        try:
            risk_analysis = {}
            
            if equity_curve_df.empty:
                return risk_analysis
            
            equity_values = equity_curve_df['equity']
            returns = equity_values.pct_change().dropna()
            
            if len(returns) == 0:
                return risk_analysis
            
            # VaR calculations
            risk_analysis['var_95'] = np.percentile(returns, 5) * 100
            risk_analysis['var_99'] = np.percentile(returns, 1) * 100
            
            # Conditional VaR (Expected Shortfall)
            var_95_threshold = np.percentile(returns, 5)
            var_99_threshold = np.percentile(returns, 1)
            
            risk_analysis['cvar_95'] = returns[returns <= var_95_threshold].mean() * 100
            risk_analysis['cvar_99'] = returns[returns <= var_99_threshold].mean() * 100
            
            # Distribution analysis
            from scipy import stats as scipy_stats
            risk_analysis['skewness'] = scipy_stats.skew(returns)
            risk_analysis['kurtosis'] = scipy_stats.kurtosis(returns)
            
            # Normality test
            _, p_value = scipy_stats.jarque_bera(returns)
            risk_analysis['normality_test_p_value'] = p_value
            risk_analysis['is_normal_distribution'] = p_value > 0.05
            
            # Downside risk metrics
            negative_returns = returns[returns < 0]
            if len(negative_returns) > 0:
                risk_analysis['downside_frequency'] = len(negative_returns) / len(returns) * 100
                risk_analysis['avg_down_day'] = negative_returns.mean() * 100
                risk_analysis['worst_day'] = negative_returns.min() * 100
            
            # Positive returns
            positive_returns = returns[returns > 0]
            if len(positive_returns) > 0:
                risk_analysis['upside_frequency'] = len(positive_returns) / len(returns) * 100
                risk_analysis['avg_up_day'] = positive_returns.mean() * 100
                risk_analysis['best_day'] = positive_returns.max() * 100
            
            # Ulcer Index
            peak = equity_values.expanding().max()
            drawdown = (equity_values - peak) / peak * 100
            risk_analysis['ulcer_index'] = np.sqrt(np.mean(drawdown ** 2))
            
            return risk_analysis
            
        except Exception as e:
            logger.error(f"Error calculating risk analysis: {e}")
            return {}
    
    def _calculate_period_analysis(self, trades_df: pd.DataFrame) -> Dict[str, Any]:
        """Calculate monthly and quarterly performance breakdown"""
        try:
            period_analysis = {}
            
            if trades_df.empty:
                return period_analysis
            
            trades_df['timestamp'] = pd.to_datetime(trades_df['timestamp'])
            trades_df['year'] = trades_df['timestamp'].dt.year
            trades_df['month'] = trades_df['timestamp'].dt.month
            trades_df['quarter'] = trades_df['timestamp'].dt.quarter
            
            # Monthly analysis
            monthly_stats = trades_df.groupby(['year', 'month']).agg({
                'net_pnl': ['sum', 'count', 'mean'],
                'timestamp': ['min', 'max']
            }).round(2)
            
            monthly_stats.columns = ['total_pnl', 'trade_count', 'avg_pnl', 'first_trade', 'last_trade']
            monthly_stats = monthly_stats.reset_index()
            monthly_stats['month_year'] = monthly_stats['year'].astype(str) + '-' + monthly_stats['month'].astype(str).str.zfill(2)
            monthly_stats['win_rate'] = (monthly_stats['total_pnl'] > 0).astype(int)
            
            period_analysis['monthly'] = monthly_stats.to_dict('records')
            
            # Quarterly analysis
            quarterly_stats = trades_df.groupby(['year', 'quarter']).agg({
                'net_pnl': ['sum', 'count', 'mean'],
                'timestamp': ['min', 'max']
            }).round(2)
            
            quarterly_stats.columns = ['total_pnl', 'trade_count', 'avg_pnl', 'first_trade', 'last_trade']
            quarterly_stats = quarterly_stats.reset_index()
            quarterly_stats['quarter_year'] = quarterly_stats['year'].astype(str) + '-Q' + quarterly_stats['quarter'].astype(str)
            quarterly_stats['win_rate'] = (quarterly_stats['total_pnl'] > 0).astype(int)
            
            period_analysis['quarterly'] = quarterly_stats.to_dict('records')
            
            # Summary statistics
            period_analysis['summary'] = {
                'profitable_months': len(monthly_stats[monthly_stats['total_pnl'] > 0]),
                'total_months': len(monthly_stats),
                'monthly_win_rate': len(monthly_stats[monthly_stats['total_pnl'] > 0]) / len(monthly_stats) * 100 if len(monthly_stats) > 0 else 0,
                'best_month_pnl': monthly_stats['total_pnl'].max() if len(monthly_stats) > 0 else 0,
                'worst_month_pnl': monthly_stats['total_pnl'].min() if len(monthly_stats) > 0 else 0,
                'profitable_quarters': len(quarterly_stats[quarterly_stats['total_pnl'] > 0]),
                'total_quarters': len(quarterly_stats),
                'quarterly_win_rate': len(quarterly_stats[quarterly_stats['total_pnl'] > 0]) / len(quarterly_stats) * 100 if len(quarterly_stats) > 0 else 0
            }
            
            return period_analysis
            
        except Exception as e:
            logger.error(f"Error calculating period analysis: {e}")
            return {}
    
    def _create_executive_summary(self, performance_metrics: Dict[str, Any], 
                                risk_metrics: Dict[str, Any], 
                                trade_statistics: Dict[str, Any]) -> Dict[str, Any]:
        """Create executive summary"""
        try:
            summary = {
                'key_highlights': [],
                'performance_grade': 'N/A',
                'risk_grade': 'N/A',
                'overall_recommendation': 'HOLD'
            }
            
            # Performance highlights
            total_return = performance_metrics.get('total_return_pct', 0)
            if total_return > 0:
                summary['key_highlights'].append(f"Generated positive returns of {total_return:.2f}%")
            else:
                summary['key_highlights'].append(f"Negative returns of {total_return:.2f}%")
            
            sharpe_ratio = performance_metrics.get('sharpe_ratio', 0)
            if sharpe_ratio > 1.0:
                summary['key_highlights'].append(f"Strong risk-adjusted returns (Sharpe: {sharpe_ratio:.2f})")
            elif sharpe_ratio > 0.5:
                summary['key_highlights'].append(f"Moderate risk-adjusted returns (Sharpe: {sharpe_ratio:.2f})")
            else:
                summary['key_highlights'].append(f"Poor risk-adjusted returns (Sharpe: {sharpe_ratio:.2f})")
            
            win_rate = performance_metrics.get('win_rate_pct', 0)
            if win_rate >= 60:
                summary['key_highlights'].append(f"High win rate of {win_rate:.1f}%")
            elif win_rate >= 45:
                summary['key_highlights'].append(f"Moderate win rate of {win_rate:.1f}%")
            else:
                summary['key_highlights'].append(f"Low win rate of {win_rate:.1f}%")
            
            max_drawdown = performance_metrics.get('max_drawdown_pct', 0)
            if max_drawdown <= 10:
                summary['key_highlights'].append(f"Low maximum drawdown of {max_drawdown:.1f}%")
            elif max_drawdown <= 20:
                summary['key_highlights'].append(f"Moderate maximum drawdown of {max_drawdown:.1f}%")
            else:
                summary['key_highlights'].append(f"High maximum drawdown of {max_drawdown:.1f}%")
            
            # Performance grade
            if total_return > 15 and sharpe_ratio > 1.0:
                summary['performance_grade'] = 'A'
            elif total_return > 8 and sharpe_ratio > 0.7:
                summary['performance_grade'] = 'B'
            elif total_return > 0 and sharpe_ratio > 0.3:
                summary['performance_grade'] = 'C'
            else:
                summary['performance_grade'] = 'D'
            
            # Risk grade
            if max_drawdown < 10 and risk_metrics.get('var_95', 0) > -2:
                summary['risk_grade'] = 'A'
            elif max_drawdown < 15 and risk_metrics.get('var_95', 0) > -3:
                summary['risk_grade'] = 'B'
            elif max_drawdown < 25:
                summary['risk_grade'] = 'C'
            else:
                summary['risk_grade'] = 'D'
            
            # Overall recommendation
            if summary['performance_grade'] in ['A', 'B'] and summary['risk_grade'] in ['A', 'B']:
                summary['overall_recommendation'] = 'BUY'
            elif summary['performance_grade'] == 'C' or summary['risk_grade'] == 'C':
                summary['overall_recommendation'] = 'HOLD'
            else:
                summary['overall_recommendation'] = 'AVOID'
            
            return summary
            
        except Exception as e:
            logger.error(f"Error creating executive summary: {e}")
            return {}
    
    def _generate_recommendations(self, performance_metrics: Dict[str, Any],
                                risk_metrics: Dict[str, Any],
                                optimization_insights: Dict[str, Any]) -> List[str]:
        """Generate actionable recommendations"""
        try:
            recommendations = []
            
            # Performance-based recommendations
            sharpe_ratio = performance_metrics.get('sharpe_ratio', 0)
            if sharpe_ratio < 1.0:
                recommendations.append("Consider improving risk-adjusted returns by optimizing position sizing or entry/exit criteria")
            
            max_drawdown = performance_metrics.get('max_drawdown_pct', 0)
            if max_drawdown > 20:
                recommendations.append("High drawdown detected - implement stricter risk management and stop-loss mechanisms")
            
            win_rate = performance_metrics.get('win_rate_pct', 0)
            if win_rate < 45:
                recommendations.append("Low win rate suggests need for better trade selection or signal filtering")
            
            profit_factor = performance_metrics.get('profit_factor', 0)
            if profit_factor < 1.2:
                recommendations.append("Profit factor below 1.2 indicates need for better trade management")
            
            # Risk-based recommendations
            var_95 = risk_metrics.get('var_95', 0)
            if var_95 < -5:
                recommendations.append("High Value-at-Risk suggests implementing position size limits")
            
            # Optimization-based recommendations
            if optimization_insights:
                if 'parameter_importance' in optimization_insights:
                    high_impact_params = [k for k, v in optimization_insights['parameter_importance'].items() if v > 0.3]
                    if high_impact_params:
                        recommendations.append(f"Focus optimization efforts on high-impact parameters: {', '.join(high_impact_params)}")
            
            # General recommendations
            if not recommendations:
                recommendations.append("Strategy shows acceptable performance - monitor for regime changes")
            
            return recommendations
            
        except Exception as e:
            logger.error(f"Error generating recommendations: {e}")
            return ["Error generating recommendations"]
    
    def _prepare_chart_data(self, trades_df: pd.DataFrame, equity_curve_df: pd.DataFrame,
                          market_data: pd.DataFrame = None) -> Dict[str, Any]:
        """Prepare data for charts and visualizations"""
        try:
            charts = {}
            
            # Equity curve data
            if not equity_curve_df.empty:
                equity_curve_df['timestamp'] = pd.to_datetime(equity_curve_df['timestamp'])
                charts['equity_curve'] = {
                    'x': equity_curve_df['timestamp'].dt.strftime('%Y-%m-%d').tolist(),
                    'y': equity_curve_df['equity'].tolist()
                }
                
                # Drawdown data
                peak = equity_curve_df['equity'].expanding().max()
                drawdown = (equity_curve_df['equity'] - peak) / peak * 100
                charts['drawdown'] = {
                    'x': equity_curve_df['timestamp'].dt.strftime('%Y-%m-%d').tolist(),
                    'y': drawdown.tolist()
                }
            
            # Trade distribution
            if not trades_df.empty:
                charts['pnl_distribution'] = trades_df['net_pnl'].tolist()
                
                # Monthly returns
                trades_df['timestamp'] = pd.to_datetime(trades_df['timestamp'])
                monthly_pnl = trades_df.groupby(trades_df['timestamp'].dt.to_period('M'))['net_pnl'].sum()
                charts['monthly_returns'] = {
                    'x': [str(period) for period in monthly_pnl.index],
                    'y': monthly_pnl.tolist()
                }
            
            return charts
            
        except Exception as e:
            logger.error(f"Error preparing chart data: {e}")
            return {}
    
    def _generate_html_report(self, report_data: Dict[str, Any]) -> str:
        """Generate HTML report"""
        try:
            logger.info("Generating HTML report...")
            
            # Create HTML template if it doesn't exist
            template_path = Path(self.template_dir) / "report_template.html"
            if not template_path.exists():
                self._create_default_html_template()
            
            # Load template
            template = self.jinja_env.get_template("report_template.html")
            
            # Render HTML
            html_content = template.render(**report_data, colors=self.colors)
            
            # Save HTML file
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            html_filename = f"backtest_report_{timestamp}.html"
            html_path = Path(self.output_dir) / html_filename
            
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            logger.info(f"HTML report saved to {html_path}")
            return str(html_path)
            
        except Exception as e:
            logger.error(f"Error generating HTML report: {e}")
            return ""
    
    def _generate_pdf_report(self, report_data: Dict[str, Any]) -> str:
        """Generate PDF report"""
        try:
            if not WEASYPRINT_AVAILABLE:
                logger.warning("WeasyPrint not available - cannot generate PDF")
                return ""
            
            logger.info("Generating PDF report...")
            
            # First generate HTML
            html_path = self._generate_html_report(report_data)
            
            if not html_path:
                return ""
            
            # Convert HTML to PDF
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            pdf_filename = f"backtest_report_{timestamp}.pdf"
            pdf_path = Path(self.output_dir) / pdf_filename
            
            weasyprint.HTML(filename=html_path).write_pdf(str(pdf_path))
            
            logger.info(f"PDF report saved to {pdf_path}")
            return str(pdf_path)
            
        except Exception as e:
            logger.error(f"Error generating PDF report: {e}")
            return ""
    
    def _generate_excel_report(self, report_data: Dict[str, Any]) -> str:
        """Generate Excel report with multiple sheets"""
        try:
            if not OPENPYXL_AVAILABLE:
                logger.warning("openpyxl not available - cannot generate Excel report")
                return ""
            
            logger.info("Generating Excel report...")
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_filename = f"backtest_report_{timestamp}.xlsx"
            excel_path = Path(self.output_dir) / excel_filename
            
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Executive Summary sheet
            self._create_excel_summary_sheet(wb, report_data)
            
            # Performance Metrics sheet
            self._create_excel_performance_sheet(wb, report_data)
            
            # Trade Details sheet
            self._create_excel_trades_sheet(wb, report_data)
            
            # Risk Analysis sheet
            self._create_excel_risk_sheet(wb, report_data)
            
            # Charts sheet (if matplotlib available)
            if MATPLOTLIB_AVAILABLE:
                self._create_excel_charts_sheet(wb, report_data)
            
            # Save workbook
            wb.save(excel_path)
            
            logger.info(f"Excel report saved to {excel_path}")
            return str(excel_path)
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return ""
    
    def _create_default_html_template(self):
        """Create default HTML template"""
        try:
            template_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ metadata.report_title }}</title>
    <style>
        body {
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            margin: 0;
            padding: 20px;
            background-color: #f8f9fa;
            color: #333;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            padding: 30px;
            border-radius: 10px;
            box-shadow: 0 0 20px rgba(0,0,0,0.1);
        }
        .header {
            text-align: center;
            border-bottom: 3px solid {{ colors.primary }};
            padding-bottom: 20px;
            margin-bottom: 30px;
        }
        .header h1 {
            color: {{ colors.primary }};
            margin: 0;
            font-size: 2.5em;
        }
        .header h2 {
            color: {{ colors.secondary }};
            margin: 10px 0;
        }
        .section {
            margin-bottom: 40px;
        }
        .section h3 {
            color: {{ colors.primary }};
            border-bottom: 2px solid {{ colors.light }};
            padding-bottom: 10px;
        }
        .metrics-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin: 20px 0;
        }
        .metric-card {
            background: {{ colors.light }};
            padding: 20px;
            border-radius: 8px;
            border-left: 4px solid {{ colors.primary }};
        }
        .metric-card h4 {
            margin: 0 0 10px 0;
            color: {{ colors.dark }};
        }
        .metric-value {
            font-size: 1.5em;
            font-weight: bold;
            color: {{ colors.primary }};
        }
        .summary-grade {
            display: inline-block;
            padding: 5px 15px;
            border-radius: 20px;
            color: white;
            font-weight: bold;
            margin: 0 10px;
        }
        .grade-a { background-color: {{ colors.success }}; }
        .grade-b { background-color: {{ colors.info }}; }
        .grade-c { background-color: {{ colors.warning }}; }
        .grade-d { background-color: {{ colors.danger }}; }
        .recommendation {
            background: {{ colors.light }};
            padding: 15px;
            border-radius: 8px;
            margin: 10px 0;
            border-left: 4px solid {{ colors.warning }};
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin: 20px 0;
        }
        th, td {
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #ddd;
        }
        th {
            background-color: {{ colors.primary }};
            color: white;
        }
        .positive { color: {{ colors.success }}; }
        .negative { color: {{ colors.danger }}; }
        .footer {
            text-align: center;
            margin-top: 40px;
            padding-top: 20px;
            border-top: 1px solid #eee;
            color: #666;
        }
    </style>
</head>
<body>
    <div class="container">
        <!-- Header -->
        <div class="header">
            <h1>{{ metadata.report_title }}</h1>
            <h2>{{ metadata.strategy_name }} - {{ metadata.symbol }}</h2>
            <p><strong>{{ metadata.company_name }}</strong></p>
            <p>Test Period: {{ metadata.test_period }}</p>
            <p>Generated: {{ metadata.generated_at }}</p>
        </div>
        
        <!-- Executive Summary -->
        <div class="section">
            <h3>Executive Summary</h3>
            <p><strong>Performance Grade:</strong> 
                <span class="summary-grade grade-{{ executive_summary.performance_grade.lower() }}">
                    {{ executive_summary.performance_grade }}
                </span>
            </p>
            <p><strong>Risk Grade:</strong> 
                <span class="summary-grade grade-{{ executive_summary.risk_grade.lower() }}">
                    {{ executive_summary.risk_grade }}
                </span>
            </p>
            <p><strong>Overall Recommendation:</strong> 
                <span class="metric-value">{{ executive_summary.overall_recommendation }}</span>
            </p>
            
            <h4>Key Highlights:</h4>
            <ul>
                {% for highlight in executive_summary.key_highlights %}
                <li>{{ highlight }}</li>
                {% endfor %}
            </ul>
        </div>
        
        <!-- Performance Metrics -->
        <div class="section">
            <h3>Performance Metrics</h3>
            <div class="metrics-grid">
                <div class="metric-card">
                    <h4>Total Return</h4>
                    <div class="metric-value {{ 'positive' if performance_metrics.total_return > 0 else 'negative' }}">
                        {{ "%.2f"|format(performance_metrics.total_return_pct or 0) }}%
                    </div>
                </div>
                <div class="metric-card">
                    <h4>Sharpe Ratio</h4>
                    <div class="metric-value">{{ "%.3f"|format(performance_metrics.sharpe_ratio or 0) }}</div>
                </div>
                <div class="metric-card">
                    <h4>Maximum Drawdown</h4>
                    <div class="metric-value negative">{{ "%.2f"|format(performance_metrics.max_drawdown_pct or 0) }}%</div>
                </div>
                <div class="metric-card">
                    <h4>Win Rate</h4>
                    <div class="metric-value">{{ "%.1f"|format(performance_metrics.win_rate_pct or 0) }}%</div>
                </div>
                <div class="metric-card">
                    <h4>Profit Factor</h4>
                    <div class="metric-value">{{ "%.2f"|format(performance_metrics.profit_factor or 0) }}</div>
                </div>
                <div class="metric-card">
                    <h4>Total Trades</h4>
                    <div class="metric-value">{{ performance_metrics.total_trades or 0 }}</div>
                </div>
            </div>
        </div>
        
        <!-- Recommendations -->
        <div class="section">
            <h3>Recommendations</h3>
            {% for recommendation in recommendations %}
            <div class="recommendation">
                {{ recommendation }}
            </div>
            {% endfor %}
        </div>
        
        <!-- Footer -->
        <div class="footer">
            <p>This report was generated by {{ metadata.author }} using advanced backtesting analysis.</p>
            <p><em>Past performance does not guarantee future results. All trading involves risk.</em></p>
        </div>
    </div>
</body>
</html>
            """
            
            template_path = Path(self.template_dir) / "report_template.html"
            with open(template_path, 'w', encoding='utf-8') as f:
                f.write(template_content.strip())
            
            logger.info(f"Default HTML template created at {template_path}")
            
        except Exception as e:
            logger.error(f"Error creating default HTML template: {e}")
    
    def _create_excel_summary_sheet(self, wb: Workbook, report_data: Dict[str, Any]):
        """Create executive summary sheet in Excel"""
        try:
            ws = wb.create_sheet("Executive Summary")
            
            # Title
            ws.merge_cells('A1:D1')
            ws['A1'] = report_data['metadata']['report_title']
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Strategy info
            ws['A3'] = "Strategy:"
            ws['B3'] = report_data['metadata']['strategy_name']
            ws['A4'] = "Symbol:"
            ws['B4'] = report_data['metadata']['symbol']
            ws['A5'] = "Test Period:"
            ws['B5'] = report_data['metadata']['test_period']
            
            # Grades
            ws['A7'] = "Performance Grade:"
            ws['B7'] = report_data['executive_summary']['performance_grade']
            ws['A8'] = "Risk Grade:"
            ws['B8'] = report_data['executive_summary']['risk_grade']
            ws['A9'] = "Overall Recommendation:"
            ws['B9'] = report_data['executive_summary']['overall_recommendation']
            
            # Key highlights
            ws['A11'] = "Key Highlights:"
            ws['A11'].font = Font(bold=True)
            
            for i, highlight in enumerate(report_data['executive_summary']['key_highlights']):
                ws[f'A{12+i}'] = f"• {highlight}"
            
        except Exception as e:
            logger.error(f"Error creating Excel summary sheet: {e}")
    
    def _create_excel_performance_sheet(self, wb: Workbook, report_data: Dict[str, Any]):
        """Create performance metrics sheet in Excel"""
        try:
            ws = wb.create_sheet("Performance Metrics")
            
            # Headers
            ws['A1'] = "Metric"
            ws['B1'] = "Value"
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            
            # Add metrics
            metrics = report_data['performance_metrics']
            row = 2
            
            metric_mappings = {
                'total_return_pct': 'Total Return (%)',
                'sharpe_ratio': 'Sharpe Ratio',
                'max_drawdown_pct': 'Maximum Drawdown (%)',
                'win_rate_pct': 'Win Rate (%)',
                'profit_factor': 'Profit Factor',
                'total_trades': 'Total Trades',
                'volatility_pct': 'Volatility (%)'
            }
            
            for key, display_name in metric_mappings.items():
                if key in metrics:
                    ws[f'A{row}'] = display_name
                    ws[f'B{row}'] = metrics[key]
                    row += 1
            
        except Exception as e:
            logger.error(f"Error creating Excel performance sheet: {e}")
    
    def _create_excel_trades_sheet(self, wb: Workbook, report_data: Dict[str, Any]):
        """Create trades detail sheet in Excel"""
        try:
            ws = wb.create_sheet("Trade Details")
            
            trades_data = report_data['trades_data']
            if not trades_data:
                ws['A1'] = "No trade data available"
                return
            
            # Create DataFrame and write to Excel
            trades_df = pd.DataFrame(trades_data)
            
            # Write headers
            for col_num, column_title in enumerate(trades_df.columns, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = column_title
                cell.font = Font(bold=True)
            
            # Write data
            for row_num, row_data in enumerate(trades_df.values, 2):
                for col_num, value in enumerate(row_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
            
        except Exception as e:
            logger.error(f"Error creating Excel trades sheet: {e}")
    
    def _create_excel_risk_sheet(self, wb: Workbook, report_data: Dict[str, Any]):
        """Create risk analysis sheet in Excel"""
        try:
            ws = wb.create_sheet("Risk Analysis")
            
            # Headers
            ws['A1'] = "Risk Metric"
            ws['B1'] = "Value"
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)
            
            # Add risk metrics
            risk_metrics = report_data['risk_analysis']
            row = 2
            
            for key, value in risk_metrics.items():
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = value
                row += 1
            
        except Exception as e:
            logger.error(f"Error creating Excel risk sheet: {e}")
    
    def _create_excel_charts_sheet(self, wb: Workbook, report_data: Dict[str, Any]):
        """Create charts sheet in Excel"""
        try:
            ws = wb.create_sheet("Charts")
            
            # Create simple chart placeholders
            ws['A1'] = "Charts and Visualizations"
            ws['A1'].font = Font(size=14, bold=True)
            
            ws['A3'] = "Note: For detailed charts, please refer to the HTML or PDF version of this report."
            
        except Exception as e:
            logger.error(f"Error creating Excel charts sheet: {e}")


# Factory function
def create_backtest_report_generator(config: Dict[str, Any] = None) -> BacktestReportGenerator:
    """Create and return BacktestReportGenerator instance"""
    return BacktestReportGenerator(config)